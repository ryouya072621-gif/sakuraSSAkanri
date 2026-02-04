import os
import json
import uuid
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, session
from openpyxl import load_workbook
from app import db
from app.models import WorkRecord, CategoryKeyword, DisplayCategory, CategoryMapping

bp = Blueprint('upload', __name__, url_prefix='/upload')


@bp.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('ファイルが選択されていません', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('ファイルが選択されていません', 'error')
            return redirect(request.url)

        if not file.filename.endswith('.xlsx'):
            flash('Excelファイル(.xlsx)を選択してください', 'error')
            return redirect(request.url)

        try:
            # Excelファイルを一時保存
            upload_id = str(uuid.uuid4())
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}.xlsx')
            file.save(filepath)

            # Excelを解析してデータを抽出（DBには保存しない）
            records, unique_combinations = parse_excel_for_preview(filepath)

            # 一時ファイルとしてJSONに保存（セッションサイズ制限回避）
            temp_data_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}_data.json')
            with open(temp_data_path, 'w', encoding='utf-8') as f:
                json.dump({
                    'records': records,
                    'unique_combinations': unique_combinations
                }, f, ensure_ascii=False, default=str)

            # セッションにはIDのみ保存
            session['upload_id'] = upload_id
            session['original_filename'] = file.filename

            # プレビュー画面へリダイレクト
            return redirect(url_for('upload.preview'))

        except Exception as e:
            flash(f'エラーが発生しました: {str(e)}', 'error')
            return redirect(request.url)

    record_count = WorkRecord.query.count()
    return render_template('upload.html', record_count=record_count)


def parse_excel_for_preview(filepath):
    """Excelファイルを解析してプレビュー用データを抽出（DBには保存しない）"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    records = []
    unique_combinations = {}

    for sheet_name in wb.sheetnames:
        if '月請求' not in sheet_name:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not row[0]:
                continue

            work_date = row[0]
            if isinstance(work_date, datetime):
                work_date = work_date.strftime('%Y-%m-%d')
            elif isinstance(work_date, str):
                try:
                    datetime.strptime(work_date, '%Y-%m-%d')
                except ValueError:
                    continue

            category1 = str(row[3]) if row[3] else ''
            category2 = str(row[4]) if row[4] else ''
            work_name = str(row[5]) if row[5] else ''

            record = {
                'work_date': work_date,
                'staff_name': str(row[1]) if row[1] else '',
                'department': str(row[2]) if row[2] else '',
                'category1': category1,
                'category2': category2,
                'work_name': work_name,
                'unit_price': int(row[6]) if row[6] else 0,
                'quantity': int(row[7]) if row[7] else 0,
                'total_amount': int(row[8]) if row[8] else 0,
                'status': str(row[9]) if row[9] else '',
                'source_month': sheet_name
            }
            records.append(record)

            # ユニーク組み合わせを抽出
            combo_key = f"{category1}|{category2}|{work_name}"
            if combo_key not in unique_combinations:
                unique_combinations[combo_key] = {
                    'category1': category1,
                    'category2': category2,
                    'work_name': work_name,
                    'count': 0
                }
            unique_combinations[combo_key]['count'] += 1

    wb.close()
    return records, list(unique_combinations.values())


def process_excel(filepath):
    """Excelファイルを処理してDBに登録"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    total_count = 0
    all_dates = []

    # まず全シートから日付を収集
    for sheet_name in wb.sheetnames:
        if '月請求' not in sheet_name:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row[0]:
                continue
            work_date = row[0]
            if isinstance(work_date, datetime):
                all_dates.append(work_date.date())
            elif isinstance(work_date, str):
                try:
                    all_dates.append(datetime.strptime(work_date, '%Y-%m-%d').date())
                except ValueError:
                    continue

    # 該当期間のデータのみ削除（他の月のデータは保持）
    if all_dates:
        min_date = min(all_dates)
        max_date = max(all_dates)
        WorkRecord.query.filter(
            WorkRecord.work_date >= min_date,
            WorkRecord.work_date <= max_date
        ).delete(synchronize_session=False)
        db.session.commit()

    # ワークブックを再度開く
    wb.close()
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # 月次シートを処理（SSA〇月請求）
    for sheet_name in wb.sheetnames:
        if '月請求' not in sheet_name:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        batch = []
        for row in rows:
            if not row[0]:  # 請求日が空なら終了
                continue

            work_date = row[0]
            if isinstance(work_date, datetime):
                work_date = work_date.date()
            elif isinstance(work_date, str):
                try:
                    work_date = datetime.strptime(work_date, '%Y-%m-%d').date()
                except ValueError:
                    continue

            record = WorkRecord(
                work_date=work_date,
                staff_name=str(row[1]) if row[1] else '',
                department=str(row[2]) if row[2] else '',
                category1=str(row[3]) if row[3] else '',
                category2=str(row[4]) if row[4] else '',
                work_name=str(row[5]) if row[5] else '',
                unit_price=int(row[6]) if row[6] else 0,
                quantity=int(row[7]) if row[7] else 0,
                total_amount=int(row[8]) if row[8] else 0,
                status=str(row[9]) if row[9] else '',
                source_month=sheet_name
            )
            batch.append(record)

            if len(batch) >= 1000:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                total_count += len(batch)
                batch = []

        if batch:
            db.session.bulk_save_objects(batch)
            db.session.commit()
            total_count += len(batch)

    wb.close()
    return total_count


@bp.route('/clear', methods=['POST'])
def clear_data():
    """データをクリア"""
    WorkRecord.query.delete()
    db.session.commit()
    flash('データをクリアしました', 'success')
    return redirect(url_for('upload.index'))


@bp.route('/preview')
def preview():
    """AI分類プレビュー画面"""
    upload_id = session.get('upload_id')
    if not upload_id:
        flash('アップロードデータがありません。再度アップロードしてください。', 'error')
        return redirect(url_for('upload.index'))

    # 一時データファイルのパスを取得
    temp_data_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}_data.json')

    if not os.path.exists(temp_data_path):
        flash('アップロードデータが見つかりません。再度アップロードしてください。', 'error')
        return redirect(url_for('upload.index'))

    # データを読み込み
    with open(temp_data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    unique_combinations = data['unique_combinations']
    total_records = len(data['records'])

    # カテゴリ一覧を取得
    categories = DisplayCategory.query.order_by(DisplayCategory.sort_order).all()

    return render_template(
        'upload_preview.html',
        upload_id=upload_id,
        filename=session.get('original_filename', ''),
        total_records=total_records,
        total_combinations=len(unique_combinations),
        unique_combinations=unique_combinations,
        categories=categories
    )


@bp.route('/confirm', methods=['POST'])
def confirm():
    """確認してDBに保存"""
    upload_id = session.get('upload_id')
    if not upload_id:
        flash('アップロードデータがありません', 'error')
        return redirect(url_for('upload.index'))

    # 一時データファイルを読み込み
    temp_data_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}_data.json')
    excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}.xlsx')

    if not os.path.exists(temp_data_path):
        flash('アップロードデータが見つかりません', 'error')
        return redirect(url_for('upload.index'))

    with open(temp_data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    records = data['records']

    # POSTデータから承認された分類を取得
    approved_categories = request.get_json() or {}

    try:
        # アップロードデータの日付範囲を取得
        dates = [datetime.strptime(r['work_date'], '%Y-%m-%d').date() for r in records if r.get('work_date')]
        if dates:
            min_date = min(dates)
            max_date = max(dates)
            # 該当期間のデータのみ削除（他の月のデータは保持）
            WorkRecord.query.filter(
                WorkRecord.work_date >= min_date,
                WorkRecord.work_date <= max_date
            ).delete(synchronize_session=False)
            db.session.commit()

        # WorkRecordを保存
        batch = []
        for record in records:
            work_record = WorkRecord(
                work_date=datetime.strptime(record['work_date'], '%Y-%m-%d').date(),
                staff_name=record['staff_name'],
                department=record['department'],
                category1=record['category1'],
                category2=record['category2'],
                work_name=record['work_name'],
                unit_price=record['unit_price'],
                quantity=record['quantity'],
                total_amount=record['total_amount'],
                status=record['status'],
                source_month=record['source_month']
            )
            batch.append(work_record)

            if len(batch) >= 1000:
                db.session.bulk_save_objects(batch)
                db.session.commit()
                batch = []

        if batch:
            db.session.bulk_save_objects(batch)
            db.session.commit()

        # 承認されたAI分類をキーワードルールとして保存
        added_keywords = 0
        for combo_key, category_id in approved_categories.items():
            if category_id:
                # combo_keyからwork_nameを抽出
                parts = combo_key.split('|')
                if len(parts) >= 3:
                    work_name = parts[2]
                    if work_name:
                        # 既存のキーワードをチェック
                        existing = CategoryKeyword.query.filter_by(keyword=work_name).first()
                        if not existing:
                            keyword = CategoryKeyword(
                                keyword=work_name,
                                display_category_id=int(category_id),
                                match_type='contains',
                                priority=10,
                                is_active=True
                            )
                            db.session.add(keyword)
                            added_keywords += 1

        db.session.commit()
        CategoryMapping.clear_cache()

        # 一時ファイルを削除
        if os.path.exists(temp_data_path):
            os.remove(temp_data_path)
        if os.path.exists(excel_path):
            os.remove(excel_path)

        # セッションをクリア
        session.pop('upload_id', None)
        session.pop('original_filename', None)

        flash(f'{len(records)}件のデータを取り込みました。{added_keywords}件のキーワードルールを追加しました。', 'success')
        return {'success': True, 'records': len(records), 'keywords': added_keywords}

    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}, 500


@bp.route('/cancel', methods=['POST'])
def cancel():
    """アップロードをキャンセル"""
    upload_id = session.get('upload_id')
    if upload_id:
        # 一時ファイルを削除
        temp_data_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}_data.json')
        excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{upload_id}.xlsx')

        if os.path.exists(temp_data_path):
            os.remove(temp_data_path)
        if os.path.exists(excel_path):
            os.remove(excel_path)

        # セッションをクリア
        session.pop('upload_id', None)
        session.pop('original_filename', None)

    flash('アップロードをキャンセルしました', 'info')
    return redirect(url_for('upload.index'))
