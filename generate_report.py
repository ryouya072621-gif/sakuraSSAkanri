import sys
sys.stdout.reconfigure(encoding='utf-8')
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

wb = Workbook()

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUB_HEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUB_HEADER_FONT = Font(bold=True, color='1F4E79', size=10)
ACCENT_FILL = PatternFill('solid', fgColor='FFF2CC')
NUM_FMT = '#,##0'
NUM_FMT_1 = '#,##0.0'
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def style_header_row(ws, row, cols):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center')
    if fmt:
        cell.number_format = fmt

# ===== Sheet 1: Summary =====
ws1 = wb.active
ws1.title = 'サマリー'
ws1.sheet_properties.tabColor = '1F4E79'

ws1.merge_cells('A1:G1')
ws1['A1'] = 'AXタスク管理 x SSA実績データ 効率化効果分析レポート'
ws1['A1'].font = Font(bold=True, color='1F4E79', size=16)
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:G2')
ws1['A2'] = '分析期間: 2026年1月 | 作成日: 2026年3月16日'
ws1['A2'].font = Font(color='808080', size=10)
ws1['A2'].alignment = Alignment(horizontal='center')

# AX概要
row = 4
ws1.merge_cells(f'A{row}:C{row}')
ws1[f'A{row}'] = 'AXタスク管理表 概要'
ws1[f'A{row}'].font = Font(bold=True, color='1F4E79', size=13)

row = 5
for i, h in enumerate(['項目', '数値', '備考'], 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 3)

ax_data = [
    ('全タスク数', 93, ''),
    ('完了', 40, '達成率 75.5%（有効タスク53件中）'),
    ('断念・停止', 40, '技術的制約/要件変更等'),
    ('進行中', 9, '進行中6 + 進行中(マーケ)3'),
    ('一時停止/確認待ち', 15, ''),
    ('未着手', 3, ''),
]
for i, (item, val, note) in enumerate(ax_data):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=item)
    ws1.cell(row=r, column=2, value=val)
    ws1.cell(row=r, column=3, value=note)
    for col in range(1, 4):
        style_data_cell(ws1, r, col)
    if item == '完了':
        ws1.cell(row=r, column=2).font = Font(bold=True, color='008000', size=11)

# SSA概要
row = 13
ws1.merge_cells(f'A{row}:C{row}')
ws1[f'A{row}'] = 'SSA実績データ概要（2026年1月）'
ws1[f'A{row}'].font = Font(bold=True, color='1F4E79', size=13)

row = 14
for i, h in enumerate(['項目', '数値', '備考'], 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 3)

ssa_summary = [
    ('記録総数', '6,840件', '18部門'),
    ('実時間合計', '18,685h/月', '件数(15,175)除外後'),
    ('件数として除外', '15,175件/月', '枚数・件数・回数等'),
    ('スタッフ数', '50名以上', ''),
    ('データ期間', '2026年1月', '1ヶ月分'),
]
for i, (item, val, note) in enumerate(ssa_summary):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=item)
    ws1.cell(row=r, column=2, value=val)
    ws1.cell(row=r, column=3, value=note)
    for col in range(1, 4):
        style_data_cell(ws1, r, col)

# 削減効果
row = 21
ws1.merge_cells(f'A{row}:C{row}')
ws1[f'A{row}'] = '削減効果サマリー（AXタスク完了分）'
ws1[f'A{row}'].font = Font(bold=True, color='1F4E79', size=13)

row = 22
for i, h in enumerate(['', '低見積', '高見積'], 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 3)

effect_data = [
    ('月間削減時間', '511h', '813h', False),
    ('年間削減時間', '6,132h', '9,756h', False),
    ('月額コスト削減（時給2,000円）', '102万円', '163万円', False),
    ('年間コスト削減（時給2,000円）', '1,226万円', '1,951万円', True),
]
for i, (item, low, high, highlight) in enumerate(effect_data):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=item)
    ws1.cell(row=r, column=2, value=low)
    ws1.cell(row=r, column=3, value=high)
    for col in range(1, 4):
        style_data_cell(ws1, r, col)
    if highlight:
        for col in range(1, 4):
            ws1.cell(row=r, column=col).fill = ACCENT_FILL
        ws1.cell(row=r, column=2).font = Font(bold=True, size=12, color='C00000')
        ws1.cell(row=r, column=3).font = Font(bold=True, size=12, color='C00000')

row = 28
ws1[f'A{row}'] = '※ 件数ベース業務（エラーチェック、印刷枚数等）のRPA効果は別途。1件あたり処理時間短縮として効果発現。'
ws1[f'A{row}'].font = Font(color='808080', size=9, italic=True)
ws1[f'A{row+1}'] = '※ SSA管理アプリ/PowerBIダッシュボード等の意思決定高速化効果は定量化困難のため含まず。'
ws1[f'A{row+1}'].font = Font(color='808080', size=9, italic=True)
ws1[f'A{row+2}'] = '※ 件数ベース業務の「1件3分短縮」試算では追加で約500h/月の削減ポテンシャルあり。'
ws1[f'A{row+2}'].font = Font(color='808080', size=9, italic=True)

ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 48

# ===== Sheet 2: Dept Detail =====
ws2 = wb.create_sheet('部署別詳細')
ws2.sheet_properties.tabColor = '2E75B6'

ws2.merge_cells('A1:I1')
ws2['A1'] = 'AXタスク完了 x SSA部門別業務 詳細マッピング'
ws2['A1'].font = Font(bold=True, color='1F4E79', size=14)
ws2['A1'].alignment = Alignment(horizontal='center')

row = 3
headers = ['部署(AX)', 'SSA対応部門', 'SSA実時間/月', 'AX完了タスク', '対象SSA業務', '対象時間', '削減率見込', '削減(低)', '削減(高)']
for i, h in enumerate(headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, len(headers))

dept_mapping = [
    ['経理課', '経理部', 1654,
     '請求書作成自動化(GAS)\nカード利用明細RPA',
     'コンカー(142h)\n問合せ対応(138h+68h)\n入金(330h)\n個人経費(75h)\nアメックス(52h)',
     805, '15-25%', 121, 201],
    ['医事(カルテ)管理課', '関東事務部', 2199,
     'エラーチェックRPA\n※件数ベース(2,442件/月)',
     'メール対応(107h)\n問合せ対応(80h+44h)\nチャット対応(37h)\n※エラーチェック自体は件数',
     268, '20-35%', 54, 94],
    ['CS課', 'CS部', 1135,
     '空き枠報告自動化\n技工アプリアンケート',
     'EPARK空き枠管理(75h)\n矯正相談日報(37h)\nE-PARK予約受付(34h)',
     146, '40-55%', 58, 80],
    ['人材開発課', '人材開発', 2156,
     '会議記録AI(Zoom+LM)\n採用AI活用\nコミュニケーション補助',
     '一次受付(382h+380h)\nスカウト送信(100h+97h+132h)\nメール対応(140h)',
     1231, '10-17%', 123, 209],
    ['訪問歯科', '訪問事務', 5647,
     'IVR導入(春日井)\n訪問経費見える化',
     '電話(142h)\n各種集計(70h)\n1on1面談(56h+28h)',
     296, '15-25%', 44, 74],
    ['庶務課', '庶務部', 2224,
     '物販管理シート効率化\n自動メール送信',
     '書類振り分け(234h)\n回収/納品(248h)\n両替金管理(56h)',
     538, '8-14%', 43, 75],
    ['事業分析課', '戦略分析部', 262,
     'RPA範囲拡大\nシート自動反映\n社内報データ',
     '社内報作成(12h)\nデータ収集RPA(7h)\n技工部月報(7h)',
     26, '50-80%', 13, 21],
    ['経営管理課', '01経営管理課', 645,
     'Jグランツ検索アプリ\nClaude導入',
     '※直接的な時間削減より意思決定高速化',
     50, '10-20%', 5, 10],
    ['運営支援課', '（全部門）', 18685,
     'SSA管理アプリ開発\nPowerBI(4種)\nGoogle Chat移行',
     '全部門のデータ可視化基盤\n意思決定の高速化',
     0, '間接効果', 50, 50],
]

for i, dm in enumerate(dept_mapping):
    r = row + 1 + i
    for col_idx, val in enumerate(dm, 1):
        ws2.cell(row=r, column=col_idx, value=val)
        style_data_cell(ws2, r, col_idx)
        ws2.cell(row=r, column=col_idx).alignment = Alignment(vertical='top', wrap_text=True)
    ws2.cell(row=r, column=3).number_format = NUM_FMT
    ws2.cell(row=r, column=6).number_format = NUM_FMT
    ws2.cell(row=r, column=8).number_format = NUM_FMT
    ws2.cell(row=r, column=9).number_format = NUM_FMT

r = row + 1 + len(dept_mapping)
ws2.cell(row=r, column=1, value='合計')
ws2.cell(row=r, column=8, value=511)
ws2.cell(row=r, column=9, value=814)
for col in range(1, 10):
    ws2.cell(row=r, column=col).fill = SUB_HEADER_FILL
    ws2.cell(row=r, column=col).font = Font(bold=True)
ws2.cell(row=r, column=8).number_format = NUM_FMT
ws2.cell(row=r, column=9).number_format = NUM_FMT

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 38
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 11
ws2.column_dimensions['I'].width = 11
for i in range(len(dept_mapping)):
    ws2.row_dimensions[row + 1 + i].height = 75

# ===== Sheet 3: SSA Business List =====
ws3 = wb.create_sheet('SSA業務一覧(時間)')
ws3.sheet_properties.tabColor = '548235'

ws3.merge_cells('A1:F1')
ws3['A1'] = 'SSA 部門別業務一覧（時間ベースのみ、件数除外）'
ws3['A1'].font = Font(bold=True, color='1F4E79', size=14)

with open('report_data.json', 'r', encoding='utf-8') as f:
    rdata = json.load(f)

row = 3
for i, h in enumerate(['部門', '業務名', '時間(h)', '業務タイプ', '単位'], 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 5)

r = row + 1
for dept, d in rdata['departments'].items():
    if d['total_hours'] < 30:
        continue
    ws3.cell(row=r, column=1, value=dept)
    ws3.cell(row=r, column=3, value=d['total_hours'])
    ws3.cell(row=r, column=3).number_format = NUM_FMT
    ws3.cell(row=r, column=5, value='時間合計')
    for col in range(1, 6):
        ws3.cell(row=r, column=col).fill = SUB_HEADER_FILL
        ws3.cell(row=r, column=col).font = SUB_HEADER_FONT
    r += 1
    for item in d['top_hours']:
        ws3.cell(row=r, column=2, value=item['work_name'])
        ws3.cell(row=r, column=3, value=item['qty'])
        ws3.cell(row=r, column=3).number_format = NUM_FMT_1
        ws3.cell(row=r, column=4, value=item['work_type'])
        ws3.cell(row=r, column=5, value='時間')
        for col in range(1, 6):
            style_data_cell(ws3, r, col)
        r += 1
    r += 1

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 60
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 10

# ===== Sheet 4: Count-based =====
ws4 = wb.create_sheet('件数ベース業務')
ws4.sheet_properties.tabColor = 'BF8F00'

ws4.merge_cells('A1:E1')
ws4['A1'] = '件数ベース業務一覧（時間集計から除外した業務）'
ws4['A1'].font = Font(bold=True, color='1F4E79', size=14)

row = 3
for i, h in enumerate(['部門', '業務名', '件数/月', 'RPA化候補', '1件3分短縮で月間削減'], 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 5)

r = row + 1
for dept, d in rdata['departments'].items():
    if not d['top_counts']:
        continue
    for item in d['top_counts']:
        ws4.cell(row=r, column=1, value=dept)
        ws4.cell(row=r, column=2, value=item['work_name'])
        ws4.cell(row=r, column=3, value=item['qty'])
        ws4.cell(row=r, column=3).number_format = NUM_FMT
        rpa = 'Yes' if item['qty'] > 100 else ''
        ws4.cell(row=r, column=4, value=rpa)
        if rpa == 'Yes':
            ws4.cell(row=r, column=4).font = Font(bold=True, color='008000')
        if item['qty'] > 0:
            save_h = round(item['qty'] * 3 / 60, 1)
            ws4.cell(row=r, column=5, value=save_h)
            ws4.cell(row=r, column=5).number_format = '#,##0.0"h"'
        for col in range(1, 6):
            style_data_cell(ws4, r, col)
        r += 1

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 60
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 22

# ===== Sheet 5: AX Task List =====
ws5 = wb.create_sheet('AXタスク一覧')
ws5.sheet_properties.tabColor = '7030A0'

ax_df = pd.read_excel('C:/Users/houmo/Downloads/AXタスク管理表.xlsx', sheet_name='全タスク')

ws5.merge_cells('A1:F1')
ws5['A1'] = 'AXタスク管理表 全タスク一覧'
ws5['A1'].font = Font(bold=True, color='1F4E79', size=14)

row = 3
for i, h in enumerate(['部署', 'タスク名', 'ステータス', '優先度', '想定ツール', '目的/概要'], 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, 6)

status_colors = {
    '完了': Font(color='008000', bold=True),
    '対応不可／断念': Font(color='808080'),
    '進行中': Font(color='0070C0', bold=True),
    '一時停止／確認待ち': Font(color='BF8F00'),
    '未着手': Font(color='C00000'),
    '進行中（マーケ）': Font(color='0070C0', bold=True),
}

for i, (_, ax_row) in enumerate(ax_df.iterrows()):
    r = row + 1 + i
    ws5.cell(row=r, column=1, value=str(ax_row.get('部署', '')))
    ws5.cell(row=r, column=2, value=str(ax_row.get('タスク名', '')))
    status = str(ax_row.get('ステータス', ''))
    ws5.cell(row=r, column=3, value=status)
    if status in status_colors:
        ws5.cell(row=r, column=3).font = status_colors[status]
    ws5.cell(row=r, column=4, value=str(ax_row.get('優先度', '') if pd.notna(ax_row.get('優先度')) else ''))
    ws5.cell(row=r, column=5, value=str(ax_row.get('想定ツール/技術', '') if pd.notna(ax_row.get('想定ツール/技術')) else ''))
    purpose = str(ax_row.get('目的/概要', '') if pd.notna(ax_row.get('目的/概要')) else '')
    ws5.cell(row=r, column=6, value=purpose[:100])
    for col in range(1, 7):
        style_data_cell(ws5, r, col)
        ws5.cell(row=r, column=col).alignment = Alignment(vertical='top', wrap_text=True)

ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 42
ws5.column_dimensions['C'].width = 16
ws5.column_dimensions['D'].width = 10
ws5.column_dimensions['E'].width = 32
ws5.column_dimensions['F'].width = 60

output_path = 'C:/Users/houmo/Downloads/AXタスク_SSA効率化分析レポート.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
